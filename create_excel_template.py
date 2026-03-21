from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json

# ================== 配置题库基础信息（修正版）==================
# 1. 题目基础信息配置
question_info = {
    "question_id": "HN-LX-20200606-01",
    "province": "河南",
    "question_type": "综合分析（政务创新+政策执行+基层治理）",
    "question_text": "看到别的县县长直播带货火了之后，A市就出台政策和文件要求也必须进行县长直播带货，并且进行排名通报。对于这个现象你怎么看？",
    "dimension_config": '{"现象解读&省情适配":{"max":8,"min":0},"危害根源分析":{"max":7,"min":0},"科学决策与措施":{"max":8,"min":0},"language逻辑与岗位适配":{"max":5,"min":0},"创新思维":{"max":2,"min":0}}'
}

# 2. 关键词评分规则配置（修正字段名）
keyword_rules = [
    ["HN-LX-20200606-01", "采分", "政务创新", "数字政务,政务改革", 2],
    ["HN-LX-20200606-01", "采分", "政策执行", "政策落地,政策实施", 2],
    ["HN-LX-20200606-01", "采分", "县域经济", "县域发展,县域振兴", 2],
    ["HN-LX-20200606-01", "采分", "为民服务", "以人民为中心,惠民利民", 2],
    ["HN-LX-20200606-01", "采分", "因地制宜", "因县施策,差异化发展", 1.5],
    ["HN-LX-20200606-01", "采分", "科学决策", "调研先行,精准施策", 1.5],
    ["HN-LX-20200606-01", "采分", "实事求是", "贴合实际,务实履职", 1.5],
    ["HN-LX-20200606-01", "采分", "四域多点", "一主两副,一圈两带", 1],
    ["HN-LX-20200606-01", "采分", "全链条保障", "配套保障,全流程支撑", 1],
    ["HN-LX-20200606-01", "扣分", "一刀切", "照搬照抄,统一要求", -1.5],
    ["HN-LX-20200606-01", "扣分", "形式主义", "表演式助农,为直播而直播", -1.5],
    ["HN-LX-20200606-01", "扣分", "盲目跟风", "跟风施策,照搬经验", -1.5],
    ["HN-LX-20200606-01", "扣分", "脱离河南实际", "未结合省情,忽视县域差异", -2],
    ["HN-LX-20200606-01", "扣分", "措施空泛", "缺乏可操作性,无具体细节", -2],
    ["HN-LX-20200606-01", "加分", "分类赋能", "一县一策,分类指导", 1],
    ["HN-LX-20200606-01", "加分", "第三方评估", "专业评估,客观评估", 1],
    ["HN-LX-20200606-01", "加分", "一播多效", "直播+乡村振兴,直播+文旅", 1],
    ["HN-LX-20200606-01", "加分", "政务创新交流平台", "经验共享,比学赶超", 1]
]

# ================== 生成最终版Excel模板 ==================
def create_question_bank_excel(save_path: str = "公考面试题库模板_最终版.xlsx"):
    """生成无拼写错误、配置完整的公考面试题库Excel模板"""
    # 1. 创建工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认工作表

    # 2. 处理question_bank工作表（修正字段名）
    ws1 = wb.create_sheet(title="question_bank")
    headers1 = ["question_id", "province", "question_type", "question_text", "dimension_config"]
    ws1.append(headers1)
    # 填充修正后的题目数据
    ws1.append([
        question_info["question_id"],
        question_info["province"],
        question_info["question_type"],
        question_info["question_text"],
        question_info["dimension_config"]
    ])

    # 3. 处理keyword_rules工作表（修正字段名）
    ws2 = wb.create_sheet(title="keyword_rules")
    headers2 = ["question_id", "keyword_type", "keyword", "synonyms", "score_change"]
    ws2.append(headers2)
    # 填充关键词规则
    for rule in keyword_rules:
        ws2.append(rule)

    # 4. 美化表格（提升可读性）
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # 美化question_bank表头
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    # 调整列宽（避免内容截断）
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 38
    ws1.column_dimensions["D"].width = 85
    ws1.column_dimensions["E"].width = 70  # 加宽维度配置列，避免JSON截断

    # 美化keyword_rules表头
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    # 调整列宽
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 45
    ws2.column_dimensions["E"].width = 18

    # 5. 保存文件
    wb.save(save_path)
    print(f"✅ 最终版题库模板已生成！保存路径：{save_path}")
    print(f"✅ 修正点：1. 字段名拼写错误 2. 维度配置JSON完整 3. 列宽适配内容不截断")

# ================== 执行生成 ==================
if __name__ == "__main__":
    create_question_bank_excel()